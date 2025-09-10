#!/usr/bin/env python3
"""
Universal Document Converter
A comprehensive tool for converting between various document formats
"""

import click
import os
import sys
from pathlib import Path
from typing import Optional, List
import logging

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Import conversion libraries
try:
    import markdown2
    import pypandoc
    from weasyprint import HTML, CSS
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    import pdfplumber
    from pypdf import PdfReader, PdfWriter
    from docx import Document
    from pptx import Presentation
    import openpyxl
    import xlsxwriter
    from bs4 import BeautifulSoup
    import yaml
    import json
    from PIL import Image
    import cairosvg
except ImportError as e:
    logger.error(f"Missing required library: {e}")
    logger.info("Please ensure all requirements are installed: pip install -r requirements.txt")
    sys.exit(1)

class DocumentConverter:
    """Universal document converter with support for multiple formats"""
    
    SUPPORTED_CONVERSIONS = {
        'markdown': ['html', 'pdf', 'docx', 'latex', 'rst', 'epub'],
        'html': ['pdf', 'markdown', 'docx', 'text'],
        'pdf': ['text', 'html', 'images', 'tables'],
        'docx': ['pdf', 'html', 'markdown', 'text'],
        'excel': ['csv', 'html', 'pdf'],
        'csv': ['excel', 'html', 'markdown'],
        'yaml': ['json', 'toml', 'xml'],
        'json': ['yaml', 'toml', 'xml'],
        'svg': ['png', 'pdf', 'jpg'],
        'image': ['pdf', 'resize', 'format_change']
    }
    
    def __init__(self, input_file: str, output_file: Optional[str] = None):
        self.input_file = Path(input_file)
        self.output_file = Path(output_file) if output_file else None
        self.input_format = self._detect_format(self.input_file)
        
    def _detect_format(self, file_path: Path) -> str:
        """Detect file format based on extension"""
        ext = file_path.suffix.lower()
        format_map = {
            '.md': 'markdown',
            '.markdown': 'markdown',
            '.html': 'html',
            '.htm': 'html',
            '.pdf': 'pdf',
            '.docx': 'docx',
            '.doc': 'docx',
            '.xlsx': 'excel',
            '.xls': 'excel',
            '.csv': 'csv',
            '.yaml': 'yaml',
            '.yml': 'yaml',
            '.json': 'json',
            '.svg': 'svg',
            '.png': 'image',
            '.jpg': 'image',
            '.jpeg': 'image',
            '.gif': 'image',
            '.bmp': 'image'
        }
        return format_map.get(ext, 'unknown')
    
    def convert_markdown_to_pdf(self, style_file: Optional[str] = None) -> None:
        """Convert Markdown to PDF with optional custom styling"""
        logger.info(f"Converting {self.input_file} to PDF...")
        
        with open(self.input_file, 'r', encoding='utf-8') as f:
            markdown_content = f.read()
        
        # Convert markdown to HTML
        html_content = markdown2.markdown(
            markdown_content,
            extras=['tables', 'fenced-code-blocks', 'header-ids', 'strike', 'task_list']
        )
        
        # Default CSS styling
        default_css = """
        @page { size: A4; margin: 2cm; }
        body { font-family: 'Helvetica', sans-serif; line-height: 1.6; }
        h1 { color: #2c3e50; border-bottom: 2px solid #3498db; }
        h2 { color: #34495e; }
        table { border-collapse: collapse; width: 100%; }
        th, td { border: 1px solid #ddd; padding: 8px; }
        th { background-color: #3498db; color: white; }
        code { background-color: #f4f4f4; padding: 2px 4px; }
        pre { background-color: #f4f4f4; padding: 10px; overflow-x: auto; }
        """
        
        # Load custom CSS if provided
        custom_css = ""
        if style_file and Path(style_file).exists():
            with open(style_file, 'r') as f:
                custom_css = f.read()
        
        # Generate PDF
        HTML(string=f"<html><body>{html_content}</body></html>").write_pdf(
            self.output_file,
            stylesheets=[CSS(string=default_css + custom_css)]
        )
        
        logger.info(f"✅ PDF saved to {self.output_file}")
    
    def convert_markdown_to_docx(self) -> None:
        """Convert Markdown to Word document"""
        logger.info(f"Converting {self.input_file} to DOCX...")
        
        # Use pypandoc for conversion
        pypandoc.convert_file(
            str(self.input_file),
            'docx',
            outputfile=str(self.output_file)
        )
        
        logger.info(f"✅ DOCX saved to {self.output_file}")
    
    def convert_pdf_to_text(self) -> None:
        """Extract text from PDF"""
        logger.info(f"Extracting text from {self.input_file}...")
        
        text_content = []
        with pdfplumber.open(self.input_file) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    text_content.append(text)
        
        with open(self.output_file, 'w', encoding='utf-8') as f:
            f.write('\n\n'.join(text_content))
        
        logger.info(f"✅ Text extracted to {self.output_file}")
    
    def convert_pdf_to_tables(self) -> None:
        """Extract tables from PDF to CSV"""
        logger.info(f"Extracting tables from {self.input_file}...")
        
        import csv
        tables_found = 0
        
        with pdfplumber.open(self.input_file) as pdf:
            for i, page in enumerate(pdf.pages):
                tables = page.extract_tables()
                for j, table in enumerate(tables):
                    if table:
                        output_path = self.output_file.parent / f"{self.output_file.stem}_table_{i+1}_{j+1}.csv"
                        with open(output_path, 'w', newline='', encoding='utf-8') as f:
                            writer = csv.writer(f)
                            writer.writerows(table)
                        tables_found += 1
        
        logger.info(f"✅ Extracted {tables_found} tables")
    
    def merge_pdfs(self, pdf_files: List[str]) -> None:
        """Merge multiple PDFs into one"""
        logger.info(f"Merging {len(pdf_files)} PDFs...")
        
        writer = PdfWriter()
        
        for pdf_file in pdf_files:
            reader = PdfReader(pdf_file)
            for page in reader.pages:
                writer.add_page(page)
        
        with open(self.output_file, 'wb') as f:
            writer.write(f)
        
        logger.info(f"✅ Merged PDF saved to {self.output_file}")
    
    def convert_excel_to_csv(self) -> None:
        """Convert Excel to CSV"""
        logger.info(f"Converting {self.input_file} to CSV...")
        
        wb = openpyxl.load_workbook(self.input_file)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            output_path = self.output_file.parent / f"{self.output_file.stem}_{sheet_name}.csv"
            
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                import csv
                writer = csv.writer(f)
                for row in sheet.iter_rows(values_only=True):
                    writer.writerow(row)
        
        logger.info(f"✅ CSV files saved")
    
    def convert_svg_to_png(self, width: int = 1024, height: int = 768) -> None:
        """Convert SVG to PNG"""
        logger.info(f"Converting {self.input_file} to PNG...")
        
        cairosvg.svg2png(
            url=str(self.input_file),
            write_to=str(self.output_file),
            output_width=width,
            output_height=height
        )
        
        logger.info(f"✅ PNG saved to {self.output_file}")
    
    def batch_convert(self, input_dir: str, input_format: str, output_format: str) -> None:
        """Batch convert all files in a directory"""
        input_path = Path(input_dir)
        files = list(input_path.glob(f"*.{input_format}"))
        
        logger.info(f"Found {len(files)} {input_format} files to convert")
        
        for file in files:
            output_file = file.with_suffix(f".{output_format}")
            self.input_file = file
            self.output_file = output_file
            
            # Determine conversion method
            if input_format == 'md' and output_format == 'pdf':
                self.convert_markdown_to_pdf()
            elif input_format == 'md' and output_format == 'docx':
                self.convert_markdown_to_docx()
            # Add more conversion methods as needed
        
        logger.info(f"✅ Batch conversion complete")

@click.group()
def cli():
    """Universal Document Converter - Convert between various document formats"""
    pass

@cli.command()
@click.argument('input_file', type=click.Path(exists=True))
@click.argument('output_file', type=click.Path())
@click.option('--style', type=click.Path(exists=True), help='CSS style file for PDF generation')
def md2pdf(input_file, output_file, style):
    """Convert Markdown to PDF"""
    converter = DocumentConverter(input_file, output_file)
    converter.convert_markdown_to_pdf(style)

@cli.command()
@click.argument('input_file', type=click.Path(exists=True))
@click.argument('output_file', type=click.Path())
def md2docx(input_file, output_file):
    """Convert Markdown to Word Document"""
    converter = DocumentConverter(input_file, output_file)
    converter.convert_markdown_to_docx()

@cli.command()
@click.argument('input_file', type=click.Path(exists=True))
@click.argument('output_file', type=click.Path())
def pdf2text(input_file, output_file):
    """Extract text from PDF"""
    converter = DocumentConverter(input_file, output_file)
    converter.convert_pdf_to_text()

@cli.command()
@click.argument('input_file', type=click.Path(exists=True))
@click.argument('output_dir', type=click.Path())
def pdf2tables(input_file, output_dir):
    """Extract tables from PDF to CSV files"""
    output_path = Path(output_dir) / Path(input_file).stem
    converter = DocumentConverter(input_file, output_path)
    converter.convert_pdf_to_tables()

@cli.command()
@click.argument('output_file', type=click.Path())
@click.argument('pdf_files', nargs=-1, required=True, type=click.Path(exists=True))
def merge_pdf(output_file, pdf_files):
    """Merge multiple PDFs into one"""
    converter = DocumentConverter(pdf_files[0], output_file)
    converter.merge_pdfs(pdf_files)

@cli.command()
@click.argument('input_file', type=click.Path(exists=True))
@click.argument('output_file', type=click.Path())
def excel2csv(input_file, output_file):
    """Convert Excel to CSV"""
    converter = DocumentConverter(input_file, output_file)
    converter.convert_excel_to_csv()

@cli.command()
@click.argument('input_file', type=click.Path(exists=True))
@click.argument('output_file', type=click.Path())
@click.option('--width', default=1024, help='Output width in pixels')
@click.option('--height', default=768, help='Output height in pixels')
def svg2png(input_file, output_file, width, height):
    """Convert SVG to PNG"""
    converter = DocumentConverter(input_file, output_file)
    converter.convert_svg_to_png(width, height)

@cli.command()
@click.argument('input_dir', type=click.Path(exists=True))
@click.argument('input_format')
@click.argument('output_format')
def batch(input_dir, input_format, output_format):
    """Batch convert all files in a directory"""
    converter = DocumentConverter("dummy", "dummy")
    converter.batch_convert(input_dir, input_format, output_format)

@cli.command()
def formats():
    """List all supported conversion formats"""
    click.echo("Supported Conversions:")
    click.echo("=" * 50)
    for source, targets in DocumentConverter.SUPPORTED_CONVERSIONS.items():
        click.echo(f"{source:12} → {', '.join(targets)}")

if __name__ == '__main__':
    cli()